from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from managers import MANAGER_REFS, MANAGER_NAMES

# --- Константы стилей ---
FILL_HEADER_DARK  = PatternFill(start_color="2B4162", end_color="2B4162", fill_type="solid")
FILL_HEADER_LIGHT = PatternFill(start_color="3D5A80", end_color="3D5A80", fill_type="solid")
FILL_GREEN        = PatternFill(start_color="D7E8CA", end_color="D7E8CA", fill_type="solid")
FILL_BLUE         = PatternFill(start_color="D9E4F5", end_color="D9E4F5", fill_type="solid")
FILL_GRAY         = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")

FONT_BOLD_WHITE   = Font(bold=True, color="FFFFFF", name="Nunito", size=10)
FONT_NORMAL_BLACK = Font(name="Nunito", size=10)
ALIGN_CENTER      = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
ALIGN_LEFT        = Alignment(horizontal="left", vertical="bottom")
# --- Конец стилей ---


def create_report(data: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "CS Report"

    n = len(MANAGER_REFS)
    connections           = data.get("connections", {})
    activation_data       = data.get("activation", {})
    five_k_data           = data.get("from_five_thousand", {})
    disconnections_data   = data.get("disconnections", {})
    monthly_turnover_data = data.get("monthly_turnover", {})
    daily_turnover_data   = data.get("daily_turnover", {})
    avg_new_turnover_data = data.get("avg_new_turnover", {})
    total_turnover_data   = data.get("total_turnover", {})
    segments_data         = data.get("segments", {})
    new_revenue_data = data.get("new_revenue", {})
    total_revenue_company_data = data.get("total_revenue_company", {})

    # === ЗАГОЛОВКИ: строка 1 ===
    ws.merge_cells("A1:A2")
    ws["A1"] = "Менеджер"
    ws["A1"].fill = FILL_HEADER_DARK
    ws["A1"].font = FONT_BOLD_WHITE
    ws["A1"].alignment = ALIGN_CENTER

    ws.merge_cells("B1:B2")
    ws["B1"] = "Реф"
    ws["B1"].fill = FILL_HEADER_DARK
    ws["B1"].font = FONT_BOLD_WHITE
    ws["B1"].alignment = ALIGN_CENTER

    ws.merge_cells("C1:D1")
    ws["C1"] = "Подключения и отток"
    ws["C1"].fill = FILL_HEADER_DARK
    ws["C1"].font = FONT_BOLD_WHITE
    ws["C1"].alignment = ALIGN_CENTER

    ws.merge_cells("F1:F2")
    ws["F1"] = "Мерчи от 5000"
    ws["F1"].fill = FILL_HEADER_DARK
    ws["F1"].font = FONT_BOLD_WHITE
    ws["F1"].alignment = ALIGN_CENTER

    ws.merge_cells("G1:G2")
    ws["G1"] = "Отключения"
    ws["G1"].fill = FILL_HEADER_DARK
    ws["G1"].font = FONT_BOLD_WHITE
    ws["G1"].alignment = ALIGN_CENTER

    ws.merge_cells("H1:H2")
    ws["H1"] = "Новый оборот"
    ws["H1"].fill = FILL_HEADER_DARK
    ws["H1"].font = FONT_BOLD_WHITE
    ws["H1"].alignment = ALIGN_CENTER

    ws.merge_cells("I1:I2")
    ws["I1"] = "Ср. дневной оборот"
    ws["I1"].fill = FILL_HEADER_DARK
    ws["I1"].font = FONT_BOLD_WHITE
    ws["I1"].alignment = ALIGN_CENTER

    # НОВЫЙ столбец J
    ws.merge_cells("J1:J2")
    ws["J1"] = "Ср. оборот новых"
    ws["J1"].fill = FILL_HEADER_DARK
    ws["J1"].font = FONT_BOLD_WHITE
    ws["J1"].alignment = ALIGN_CENTER

    # Старый J → теперь K
    ws.merge_cells("K1:K2")
    ws["K1"] = "Общий оборот"
    ws["K1"].fill = FILL_HEADER_DARK
    ws["K1"].font = FONT_BOLD_WHITE
    ws["K1"].alignment = ALIGN_CENTER

    # Разделители L и M
    ws.merge_cells("L1:L2")
    ws["L1"] = ""

    ws.merge_cells("M1:M2")
    ws["M1"] = ""

    # Сегменты N1:S1
    ws.merge_cells("N1:S1")
    ws["N1"] = "Сегменты по обороту"
    ws["N1"].fill = FILL_HEADER_DARK
    ws["N1"].font = FONT_BOLD_WHITE
    ws["N1"].alignment = ALIGN_CENTER

    # === ПОДЗАГОЛОВКИ: строка 2 ===
    ws["C2"] = "Новые подкл."
    ws["C2"].fill = FILL_HEADER_LIGHT
    ws["C2"].font = FONT_BOLD_WHITE
    ws["C2"].alignment = ALIGN_CENTER

    ws["D2"] = "Активаций +%"
    ws["D2"].fill = FILL_HEADER_LIGHT
    ws["D2"].font = FONT_BOLD_WHITE
    ws["D2"].alignment = ALIGN_CENTER

    ws.merge_cells("T1:T2")
    ws["T1"] = ""  # разделитель

    ws.merge_cells("U1:V1")
    ws["U1"] = "Доходность"
    ws["U1"].fill = FILL_HEADER_DARK
    ws["U1"].font = FONT_BOLD_WHITE
    ws["U1"].alignment = ALIGN_CENTER

    ws["U2"] = "Новых мерчей"
    ws["U2"].fill = FILL_HEADER_LIGHT
    ws["U2"].font = FONT_BOLD_WHITE
    ws["U2"].alignment = ALIGN_CENTER

    ws["V2"] = "Общая за месяц"
    ws["V2"].fill = FILL_HEADER_LIGHT
    ws["V2"].font = FONT_BOLD_WHITE
    ws["V2"].alignment = ALIGN_CENTER

    segment_headers = [
        ("N", "Starter\n<$1k"),
        ("O", "Growth\n$1k-2k"),
        ("P", "Mid\n$2k-4k"),
        ("Q", "Large\n$4k-8k"),
        ("R", "VIP\n$8k-16k"),
        ("S", "Enterprise\n$16k+"),
    ]
    segment_keys = ["starter", "growth", "mid", "large", "vip", "enterprise"]
    seg_cols     = [14, 15, 16, 17, 18, 19]  # N=14 ... S=19

    for col_letter, header in segment_headers:
        ws[f"{col_letter}2"] = header
        ws[f"{col_letter}2"].fill = FILL_HEADER_LIGHT
        ws[f"{col_letter}2"].font = FONT_BOLD_WHITE
        ws[f"{col_letter}2"].alignment = ALIGN_CENTER

    # === ДАННЫЕ: строки 3..n+2 ===
    for row_idx, (ref, name) in enumerate(zip(MANAGER_REFS, MANAGER_NAMES), start=3):
        cell_a = ws.cell(row=row_idx, column=1, value=name)
        cell_a.fill = FILL_GRAY
        cell_a.font = FONT_NORMAL_BLACK
        cell_a.alignment = ALIGN_LEFT

        cell_b = ws.cell(row=row_idx, column=2, value=ref)
        cell_b.fill = FILL_HEADER_DARK
        cell_b.font = FONT_BOLD_WHITE
        cell_b.alignment = ALIGN_CENTER

        cell_c = ws.cell(row=row_idx, column=3, value=connections.get(ref, 0))
        cell_c.fill = FILL_BLUE
        cell_c.font = FONT_NORMAL_BLACK
        cell_c.alignment = ALIGN_CENTER

        cell_d = ws.cell(row=row_idx, column=4, value=activation_data.get(ref, 0.0))
        cell_d.number_format = '0.00"%"'
        cell_d.fill = FILL_BLUE
        cell_d.font = FONT_NORMAL_BLACK
        cell_d.alignment = ALIGN_CENTER

        cell_f = ws.cell(row=row_idx, column=6, value=five_k_data.get(ref, 0))
        cell_f.fill = FILL_GREEN
        cell_f.font = FONT_NORMAL_BLACK
        cell_f.alignment = ALIGN_CENTER

        cell_g = ws.cell(row=row_idx, column=7, value=disconnections_data.get(ref, 0))
        cell_g.fill = FILL_BLUE
        cell_g.font = FONT_NORMAL_BLACK
        cell_g.alignment = ALIGN_CENTER

        cell_h = ws.cell(row=row_idx, column=8, value=monthly_turnover_data.get(ref, 0.0))
        cell_h.number_format = '#,##0.00'
        cell_h.fill = FILL_GREEN
        cell_h.font = FONT_NORMAL_BLACK
        cell_h.alignment = ALIGN_CENTER

        cell_i = ws.cell(row=row_idx, column=9, value=daily_turnover_data.get(ref, 0.0))
        cell_i.number_format = '#,##0.00'
        cell_i.fill = FILL_GREEN
        cell_i.font = FONT_NORMAL_BLACK
        cell_i.alignment = ALIGN_CENTER

        # НОВЫЙ J — ср. оборот новых
        cell_j = ws.cell(row=row_idx, column=10, value=avg_new_turnover_data.get(ref, 0.0))
        cell_j.number_format = '#,##0.00'
        cell_j.fill = FILL_GREEN
        cell_j.font = FONT_NORMAL_BLACK
        cell_j.alignment = ALIGN_CENTER

        # Старый J → K — общий оборот
        cell_k = ws.cell(row=row_idx, column=11, value=total_turnover_data.get(ref, 0.0))
        cell_k.number_format = '#,##0.00'
        cell_k.fill = FILL_GREEN
        cell_k.font = FONT_NORMAL_BLACK
        cell_k.alignment = ALIGN_CENTER

        ref_segments = segments_data.get(ref, {})
        for col_num, seg_key in zip(seg_cols, segment_keys):
            cell = ws.cell(row=row_idx, column=col_num, value=ref_segments.get(seg_key, 0))
            cell.fill = FILL_BLUE
            cell.font = FONT_NORMAL_BLACK
            cell.alignment = ALIGN_CENTER

        cell_u = ws.cell(row=row_idx, column=21, value=new_revenue_data.get(ref, 0.0))
        cell_u.number_format = '#,##0.00'
        cell_u.fill = FILL_GREEN
        cell_u.font = FONT_NORMAL_BLACK
        cell_u.alignment = ALIGN_CENTER

        cell_v = ws.cell(row=row_idx, column=22, value=total_revenue_company_data.get(ref, 0.0))
        cell_v.number_format = '#,##0.00'
        cell_v.fill = FILL_GREEN
        cell_v.font = FONT_NORMAL_BLACK
        cell_v.alignment = ALIGN_CENTER

    # === ИТОГО: строка n+3 ===
    total_row = n + 3

    cell_total_c = ws.cell(row=total_row, column=3, value=sum(connections.get(ref, 0) for ref in MANAGER_REFS))
    cell_total_c.fill = FILL_HEADER_DARK
    cell_total_c.font = FONT_BOLD_WHITE
    cell_total_c.alignment = ALIGN_CENTER

    avg_act = sum(activation_data.get(ref, 0) for ref in MANAGER_REFS) / len(MANAGER_REFS)
    cell_total_d = ws.cell(row=total_row, column=4, value=avg_act)
    cell_total_d.number_format = '0.00"%"'
    cell_total_d.fill = FILL_HEADER_DARK
    cell_total_d.font = FONT_BOLD_WHITE
    cell_total_d.alignment = ALIGN_CENTER

    cell_total_f = ws.cell(row=total_row, column=6, value=sum(five_k_data.get(ref, 0) for ref in MANAGER_REFS))
    cell_total_f.fill = FILL_HEADER_DARK
    cell_total_f.font = FONT_BOLD_WHITE
    cell_total_f.alignment = ALIGN_CENTER

    cell_total_g = ws.cell(row=total_row, column=7, value=sum(disconnections_data.get(ref, 0) for ref in MANAGER_REFS))
    cell_total_g.fill = FILL_HEADER_DARK
    cell_total_g.font = FONT_BOLD_WHITE
    cell_total_g.alignment = ALIGN_CENTER

    cell_total_h = ws.cell(row=total_row, column=8,
                           value=sum(monthly_turnover_data.get(ref, 0.0) for ref in MANAGER_REFS))
    cell_total_h.number_format = '#,##0.00'
    cell_total_h.fill = FILL_HEADER_DARK
    cell_total_h.font = FONT_BOLD_WHITE
    cell_total_h.alignment = ALIGN_CENTER

    avg_daily = sum(daily_turnover_data.get(ref, 0.0) for ref in MANAGER_REFS) / len(MANAGER_REFS)
    cell_total_i = ws.cell(row=total_row, column=9, value=avg_daily)
    cell_total_i.number_format = '#,##0.00'
    cell_total_i.fill = FILL_HEADER_DARK
    cell_total_i.font = FONT_BOLD_WHITE
    cell_total_i.alignment = ALIGN_CENTER

    avg_new_total = sum(avg_new_turnover_data.get(ref, 0.0) for ref in MANAGER_REFS) / len(MANAGER_REFS)
    cell_total_j = ws.cell(row=total_row, column=10, value=avg_new_total)
    cell_total_j.number_format = '#,##0.00'
    cell_total_j.fill = FILL_HEADER_DARK
    cell_total_j.font = FONT_BOLD_WHITE
    cell_total_j.alignment = ALIGN_CENTER

    cell_total_k = ws.cell(row=total_row, column=11,
                           value=sum(total_turnover_data.get(ref, 0.0) for ref in MANAGER_REFS))
    cell_total_k.number_format = '#,##0.00'
    cell_total_k.fill = FILL_HEADER_DARK
    cell_total_k.font = FONT_BOLD_WHITE
    cell_total_k.alignment = ALIGN_CENTER

    for col_num, seg_key in zip(seg_cols, segment_keys):
        total_seg = sum(segments_data.get(ref, {}).get(seg_key, 0) for ref in MANAGER_REFS)
        cell = ws.cell(row=total_row, column=col_num, value=total_seg)
        cell.fill = FILL_HEADER_DARK
        cell.font = FONT_BOLD_WHITE
        cell.alignment = ALIGN_CENTER

    cell_total_u = ws.cell(row=total_row, column=21,
                           value=sum(new_revenue_data.get(ref, 0.0) for ref in MANAGER_REFS))
    cell_total_u.number_format = '#,##0.00'
    cell_total_u.fill = FILL_HEADER_DARK
    cell_total_u.font = FONT_BOLD_WHITE
    cell_total_u.alignment = ALIGN_CENTER

    cell_total_v = ws.cell(row=total_row, column=22,
                           value=sum(total_revenue_company_data.get(ref, 0.0) for ref in MANAGER_REFS))
    cell_total_v.number_format = '#,##0.00'
    cell_total_v.fill = FILL_HEADER_DARK
    cell_total_v.font = FONT_BOLD_WHITE
    cell_total_v.alignment = ALIGN_CENTER

    # === ШИРИНА СТОЛБЦОВ ===
    ws.column_dimensions["A"].width = 28.0
    ws.column_dimensions["B"].width = 12.0
    ws.column_dimensions["C"].width = 14.0
    ws.column_dimensions["D"].width = 18.0
    ws.column_dimensions["E"].width = 4.0
    ws.column_dimensions["F"].width = 16.0
    ws.column_dimensions["G"].width = 14.0
    ws.column_dimensions["H"].width = 18.0
    ws.column_dimensions["I"].width = 20.0
    ws.column_dimensions["J"].width = 20.0
    ws.column_dimensions["K"].width = 18.0
    ws.column_dimensions["L"].width = 4.0
    ws.column_dimensions["M"].width = 4.0
    ws.column_dimensions["N"].width = 12.0
    ws.column_dimensions["O"].width = 12.0
    ws.column_dimensions["P"].width = 12.0
    ws.column_dimensions["Q"].width = 12.0
    ws.column_dimensions["R"].width = 12.0
    ws.column_dimensions["S"].width = 14.0
    ws.column_dimensions["T"].width = 4.0
    ws.column_dimensions["U"].width = 18.0
    ws.column_dimensions["V"].width = 18.0

    return wb